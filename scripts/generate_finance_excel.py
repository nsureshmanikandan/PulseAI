"""
Generate a realistic multi-tab Finance Excel workbook for PulseAI testing.

Tabs and relationships:
  Customers      ← base entity (customer_id)
  Accounts       ← FK customer_id → Customers
  Transactions   ← FK account_id  → Accounts
  Loans          ← FK customer_id → Customers
  Loan_Payments  ← FK loan_id     → Loans
  Portfolio      ← aggregated view per customer
"""

import random
import pandas as pd
from datetime import date, timedelta
from pathlib import Path

random.seed(42)

# ── helpers ────────────────────────────────────────────────────────────────────

def rand_date(start: date, end: date) -> date:
    return start + timedelta(days=random.randint(0, (end - start).days))

def rand_dates(n, start, end):
    return [rand_date(start, end) for _ in range(n)]

START = date(2022, 1, 1)
END   = date(2024, 12, 31)

# ── 1. Customers (200 rows) ────────────────────────────────────────────────────
N_CUST = 200
FIRST  = ["Alice","Bob","Carol","David","Eva","Frank","Grace","Henry","Iris","Jake",
          "Karen","Leo","Maria","Nate","Olivia","Paul","Quinn","Rachel","Sam","Tina"]
LAST   = ["Smith","Johnson","Williams","Brown","Jones","Garcia","Miller","Davis",
          "Wilson","Taylor","Anderson","Thomas","Jackson","White","Harris","Martin"]
CITIES = ["New York","London","Singapore","Toronto","Sydney","Dubai","Mumbai","Paris","Tokyo","Berlin"]
SEGMENTS = ["Retail","Private","Corporate","Wealth","SME"]
RISK = ["Low","Medium","High"]

customer_ids = [f"C{str(i).zfill(5)}" for i in range(1, N_CUST+1)]
customers = pd.DataFrame({
    "customer_id":   customer_ids,
    "full_name":     [f"{random.choice(FIRST)} {random.choice(LAST)}" for _ in range(N_CUST)],
    "segment":       [random.choice(SEGMENTS) for _ in range(N_CUST)],
    "city":          [random.choice(CITIES) for _ in range(N_CUST)],
    "age":           [random.randint(22, 72) for _ in range(N_CUST)],
    "risk_rating":   [random.choice(RISK) for _ in range(N_CUST)],
    "onboard_date":  rand_dates(N_CUST, date(2015, 1, 1), date(2022, 12, 31)),
    "annual_income": [round(random.uniform(35_000, 350_000), 2) for _ in range(N_CUST)],
    "credit_score":  [random.randint(450, 850) for _ in range(N_CUST)],
    "active":        [random.choice(["Y", "Y", "Y", "N"]) for _ in range(N_CUST)],
})

# ── 2. Accounts (350 rows — some customers have 2 accounts) ───────────────────
N_ACC = 350
ACCT_TYPES = ["Checking","Savings","Investment","Money Market","Fixed Deposit"]
CURRENCIES = ["USD","GBP","EUR","SGD","AED"]

account_ids = [f"A{str(i).zfill(6)}" for i in range(1, N_ACC+1)]
acct_customers = random.choices(customer_ids, k=N_ACC)
accounts = pd.DataFrame({
    "account_id":    account_ids,
    "customer_id":   acct_customers,
    "account_type":  [random.choice(ACCT_TYPES) for _ in range(N_ACC)],
    "currency":      [random.choice(CURRENCIES) for _ in range(N_ACC)],
    "open_date":     rand_dates(N_ACC, date(2015, 1, 1), date(2023, 6, 30)),
    "balance":       [round(random.uniform(500, 500_000), 2) for _ in range(N_ACC)],
    "interest_rate": [round(random.uniform(0.5, 6.5), 2) for _ in range(N_ACC)],
    "status":        [random.choice(["Active","Active","Active","Dormant","Closed"]) for _ in range(N_ACC)],
    "branch":        [random.choice(["HQ","North","South","East","West","Online"]) for _ in range(N_ACC)],
})

# ── 3. Transactions (1 000 rows) ───────────────────────────────────────────────
N_TXN = 1000
TXN_TYPES = ["Deposit","Withdrawal","Transfer","Fee","Interest","Payment"]
CHANNELS   = ["Online","Branch","ATM","Mobile","Wire"]

txn_ids = [f"T{str(i).zfill(7)}" for i in range(1, N_TXN+1)]
txn_accounts = random.choices(account_ids, k=N_TXN)
txn_dates = rand_dates(N_TXN, START, END)
amounts_raw = [round(random.uniform(10, 25_000), 2) for _ in range(N_TXN)]
txn_types = [random.choice(TXN_TYPES) for _ in range(N_TXN)]

transactions = pd.DataFrame({
    "transaction_id":   txn_ids,
    "account_id":       txn_accounts,
    "transaction_date": txn_dates,
    "type":             txn_types,
    "amount":           amounts_raw,
    "debit_credit":     ["DR" if t in ("Withdrawal","Fee","Transfer") else "CR" for t in txn_types],
    "channel":          [random.choice(CHANNELS) for _ in range(N_TXN)],
    "description":      [f"{t} via {random.choice(CHANNELS)}" for t in txn_types],
    "balance_after":    [round(random.uniform(100, 600_000), 2) for _ in range(N_TXN)],
    "status":           [random.choice(["Completed","Completed","Completed","Pending","Failed"]) for _ in range(N_TXN)],
})
transactions.sort_values("transaction_date", inplace=True)

# ── 4. Loans (180 rows) ────────────────────────────────────────────────────────
N_LOAN = 180
LOAN_TYPES = ["Personal","Mortgage","Auto","Business","Education","Revolving Credit"]
loan_ids = [f"L{str(i).zfill(5)}" for i in range(1, N_LOAN+1)]
loan_customers = random.choices(customer_ids, k=N_LOAN)
disburse_dates = rand_dates(N_LOAN, date(2019, 1, 1), date(2023, 12, 31))
loan_amounts = [round(random.choice([5_000,10_000,25_000,50_000,100_000,250_000,500_000]), 2)
                for _ in range(N_LOAN)]
terms_months = [random.choice([12, 24, 36, 48, 60, 120, 240, 360]) for _ in range(N_LOAN)]
rates = [round(random.uniform(3.5, 18.5), 2) for _ in range(N_LOAN)]
paid_pct = [round(random.uniform(0, 1), 2) for _ in range(N_LOAN)]
outstanding = [round(amt * (1 - pct), 2) for amt, pct in zip(loan_amounts, paid_pct)]

loans = pd.DataFrame({
    "loan_id":           loan_ids,
    "customer_id":       loan_customers,
    "loan_type":         [random.choice(LOAN_TYPES) for _ in range(N_LOAN)],
    "disbursement_date": disburse_dates,
    "loan_amount":       loan_amounts,
    "outstanding_balance": outstanding,
    "interest_rate":     rates,
    "term_months":       terms_months,
    "monthly_emi":       [round((a * r/1200) / (1-(1+r/1200)**(-t)), 2)
                          for a, r, t in zip(loan_amounts, rates, terms_months)],
    "loan_status":       [random.choice(["Active","Active","Active","Closed","Default","Watch"])
                          for _ in range(N_LOAN)],
    "collateral":        [random.choice(["None","Property","Vehicle","Stocks","Gold"]) for _ in range(N_LOAN)],
    "dpd":               [random.choices([0,0,0,30,60,90,180], weights=[60,15,10,7,4,3,1])[0]
                          for _ in range(N_LOAN)],
})

# ── 5. Loan_Payments (600 rows) ────────────────────────────────────────────────
N_PMT = 600
pmt_loan_ids = random.choices(loan_ids, k=N_PMT)
pmt_dates = rand_dates(N_PMT, START, END)

loan_emi_map = dict(zip(loans["loan_id"], loans["monthly_emi"]))

payments = pd.DataFrame({
    "payment_id":    [f"P{str(i).zfill(7)}" for i in range(1, N_PMT+1)],
    "loan_id":       pmt_loan_ids,
    "payment_date":  pmt_dates,
    "scheduled_emi": [round(loan_emi_map.get(lid, 500), 2) for lid in pmt_loan_ids],
    "amount_paid":   [round(loan_emi_map.get(lid, 500) * random.choice([0.0, 0.5, 1.0, 1.0, 1.0, 1.2]), 2)
                      for lid in pmt_loan_ids],
    "payment_mode":  [random.choice(["Auto-Debit","Net Banking","NEFT","Cheque","Cash"]) for _ in range(N_PMT)],
    "days_overdue":  [random.choices([0,0,0,7,15,30,60], weights=[65,10,10,6,4,3,2])[0]
                      for _ in range(N_PMT)],
    "penalty_fee":   [0.0 if d == 0 else round(d * 2.5, 2) for d in
                      [random.choices([0,0,0,7,15,30,60], weights=[65,10,10,6,4,3,2])[0] for _ in range(N_PMT)]],
    "status":        [random.choice(["Paid","Paid","Paid","Partial","Missed"]) for _ in range(N_PMT)],
})
payments.sort_values("payment_date", inplace=True)

# ── 6. Portfolio (summary per customer — 200 rows) ────────────────────────────
cust_txn_map    = transactions.merge(accounts[["account_id","customer_id"]], on="account_id")
cust_txn_totals = cust_txn_map.groupby("customer_id")["amount"].sum().reset_index().rename(columns={"amount":"total_txn_value"})
cust_loan_map   = loans.groupby("customer_id").agg(
    total_loans=("loan_id","count"),
    total_loan_amount=("loan_amount","sum"),
    total_outstanding=("outstanding_balance","sum"),
).reset_index()
cust_acct_map   = accounts.groupby("customer_id").agg(
    num_accounts=("account_id","count"),
    total_balance=("balance","sum"),
).reset_index()

portfolio = customers[["customer_id","full_name","segment","risk_rating","credit_score","annual_income"]].copy()
portfolio = portfolio.merge(cust_acct_map,    on="customer_id", how="left")
portfolio = portfolio.merge(cust_loan_map,    on="customer_id", how="left")
portfolio = portfolio.merge(cust_txn_totals,  on="customer_id", how="left")

portfolio["total_loans"]        = portfolio["total_loans"].fillna(0).astype(int)
portfolio["num_accounts"]       = portfolio["num_accounts"].fillna(0).astype(int)
portfolio["total_loan_amount"]  = portfolio["total_loan_amount"].fillna(0).round(2)
portfolio["total_outstanding"]  = portfolio["total_outstanding"].fillna(0).round(2)
portfolio["total_balance"]      = portfolio["total_balance"].fillna(0).round(2)
portfolio["total_txn_value"]    = portfolio["total_txn_value"].fillna(0).round(2)
portfolio["net_worth_est"]      = (portfolio["total_balance"] - portfolio["total_outstanding"] + portfolio["annual_income"]).round(2)
portfolio["ltv_ratio"]          = (portfolio["total_outstanding"] / portfolio["annual_income"].replace(0,1)).round(3)

# ── Write Excel ────────────────────────────────────────────────────────────────
out_path = Path(r"C:\Users\n.sureshmanikandan\Repo1\PulseAI\test_data\FinanceData_MultiTab.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)

with pd.ExcelWriter(out_path, engine="openpyxl", datetime_format="YYYY-MM-DD") as writer:
    customers.to_excel(   writer, sheet_name="Customers",      index=False)
    accounts.to_excel(    writer, sheet_name="Accounts",       index=False)
    transactions.to_excel(writer, sheet_name="Transactions",   index=False)
    loans.to_excel(       writer, sheet_name="Loans",          index=False)
    payments.to_excel(    writer, sheet_name="Loan_Payments",  index=False)
    portfolio.to_excel(   writer, sheet_name="Portfolio",      index=False)

    # ── Style header rows ──────────────────────────────────────────────────────
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILLS = {
        "Customers":     "1E3A5F",
        "Accounts":      "1A4731",
        "Transactions":  "3B1F5E",
        "Loans":         "5E2A1A",
        "Loan_Payments": "1A3D5E",
        "Portfolio":     "2D4A1E",
    }

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, ws in writer.sheets.items():
        fill_color = HEADER_FILLS.get(sheet_name, "2D3748")
        header_fill = PatternFill("solid", fgColor=fill_color)
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Auto-fit columns
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 28)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Alternate row shading
        from openpyxl.styles import PatternFill as PF
        odd_fill  = PF("solid", fgColor="F8FAFC")
        even_fill = PF("solid", fgColor="EEF2FF")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_fill = odd_fill if row_idx % 2 == 0 else even_fill
            for cell in row:
                if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                    cell.fill = row_fill
                cell.border = border

print(f"[OK] Excel created: {out_path}")
print(f"   Customers:    {len(customers)} rows")
print(f"   Accounts:     {len(accounts)} rows")
print(f"   Transactions: {len(transactions)} rows")
print(f"   Loans:        {len(loans)} rows")
print(f"   Loan_Payments:{len(payments)} rows")
print(f"   Portfolio:    {len(portfolio)} rows")
print()
print("Relationships:")
print("  Accounts.customer_id     → Customers.customer_id")
print("  Transactions.account_id  → Accounts.account_id")
print("  Loans.customer_id        → Customers.customer_id")
print("  Loan_Payments.loan_id    → Loans.loan_id")
print("  Portfolio.customer_id    → Customers.customer_id (aggregated)")
