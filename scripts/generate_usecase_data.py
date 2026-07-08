"""
Generate 4 realistic multi-tab Excel datasets for PulseAI use-case demos.
All files have multiple tabs with FK relationships so the relationship finder works.

  1. Banking_LoanPortfolio.xlsx  - 5 tabs: Customers, Loans, Loan_Payments, Credit_Bureau, Collateral
  2. Insurance_Claims.xlsx       - 4 tabs: Policyholders, Policies, Claims, Fraud_Indicators
  3. Retail_Sales.xlsx           - 5 tabs: Customers, Products, Orders, Order_Items, Returns
  4. HR_Workforce.xlsx           - 5 tabs: Employees, Departments, Performance, Salary_History, Training
"""
import random
from pathlib import Path
from datetime import date, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)
OUT = Path(r"C:\Users\n.sureshmanikandan\Repo1\PulseAI\test_data")
OUT.mkdir(parents=True, exist_ok=True)


def rand_date(start="2022-01-01", end="2024-12-31"):
    s = date.fromisoformat(start)
    e = date.fromisoformat(end)
    return s + timedelta(days=random.randint(0, (e - s).days))


def style_wb(path):
    wb = load_workbook(path)
    hdr_fill  = PatternFill("solid", fgColor="0F172A")
    hdr_font  = Font(bold=True, color="60A5FA", size=10)
    alt_fill  = PatternFill("solid", fgColor="1E293B")
    norm_fill = PatternFill("solid", fgColor="0F172A")
    border    = Border(bottom=Side(style="thin", color="1E3A5F"))
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            fill = alt_fill if row[0].row % 2 == 0 else norm_fill
            for cell in row:
                cell.fill = fill
                cell.font = Font(color="CBD5E1", size=9)
                cell.alignment = Alignment(vertical="center")
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 32)
        ws.freeze_panes = "A2"
    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Banking_LoanPortfolio.xlsx
#    Tabs: Customers → Loans → Loan_Payments
#                   → Credit_Bureau
#                   → Collateral
# ═══════════════════════════════════════════════════════════════════════════════
print("Generating Banking_LoanPortfolio.xlsx ...")

CITIES    = ["Mumbai", "Delhi", "Bangalore", "Chennai", "Hyderabad", "Pune", "Kolkata", "Ahmedabad", "Jaipur", "Surat"]
SEGMENTS  = ["Retail", "SME", "Corporate", "Wealth", "Priority"]
GENDERS   = ["Male", "Female"]
OCCUPATIONS = ["Salaried", "Self-Employed", "Business Owner", "Government", "Professional"]

# --- Customers (200) ---
cust_ids   = [f"CUST{i:04d}" for i in range(1, 201)]
customers  = pd.DataFrame({
    "customer_id":    cust_ids,
    "full_name":      [f"{random.choice(['Rahul','Priya','Amit','Neha','Vijay','Sunita','Arun','Meera','Kiran','Deepak'])} {random.choice(['Sharma','Patel','Singh','Kumar','Gupta','Verma','Joshi','Nair','Reddy','Shah'])}" for _ in range(200)],
    "age":            [random.randint(22, 65) for _ in range(200)],
    "gender":         [random.choice(GENDERS) for _ in range(200)],
    "city":           [random.choice(CITIES) for _ in range(200)],
    "segment":        [random.choice(SEGMENTS) for _ in range(200)],
    "occupation":     [random.choice(OCCUPATIONS) for _ in range(200)],
    "annual_income":  [round(random.uniform(300000, 5000000), -3) for _ in range(200)],
    "phone":          [f"+91 9{random.randint(100000000, 999999999)}" for _ in range(200)],
    "email":          [f"customer{i}@email.com" for i in range(1, 201)],
    "kyc_status":     [random.choice(["Verified", "Verified", "Verified", "Pending", "Expired"]) for _ in range(200)],
    "onboarding_date":[rand_date("2015-01-01", "2023-12-31") for _ in range(200)],
})

# --- Loans (300) ---
LOAN_TYPES  = ["Personal", "Mortgage", "Auto", "Business", "Education", "Gold", "Revolving Credit"]
LOAN_STATUS = (["Active"]*52 + ["Closed"]*18 + ["Default"]*14 + ["Watch"]*10 + ["Restructured"]*6)

loan_ids    = [f"LOAN{i:05d}" for i in range(1, 301)]
l_cust_ids  = [random.choice(cust_ids) for _ in range(300)]
l_types     = [random.choice(LOAN_TYPES) for _ in range(300)]
principal   = [round(random.uniform(50000, 10000000), -2) for _ in range(300)]
int_rate    = [round(random.uniform(6.5, 18.5), 2) for _ in range(300)]
tenure      = [random.choice([12, 24, 36, 48, 60, 84, 120, 180, 240]) for _ in range(300)]
outstanding = [round(p * random.uniform(0.05, 0.95), -2) for p in principal]
emi         = [round(o / max(t * 0.75, 1), 2) for o, t in zip(outstanding, tenure)]
dpd_list    = []
statuses    = []
for _ in range(300):
    s = random.choice(LOAN_STATUS)
    if s == "Default":   d = random.randint(90, 730)
    elif s == "Watch":   d = random.randint(31, 89)
    elif s == "Active":  d = random.randint(0, 30)
    else:                d = 0
    dpd_list.append(d)
    statuses.append("Default" if d >= 90 else "Watch" if d >= 31 else s)

credit_scores = [random.randint(450, 850) for _ in range(300)]
risk_grades   = ["A" if cs >= 750 else "B" if cs >= 650 else "C" if cs >= 550 else "D" for cs in credit_scores]
disbursal     = [rand_date("2018-01-01", "2023-12-31") for _ in range(300)]
maturity      = [d + timedelta(days=t * 30) for d, t in zip(disbursal, tenure)]
branch_codes  = [f"BR{random.randint(100,999)}" for _ in range(300)]

loans = pd.DataFrame({
    "loan_id":          loan_ids,
    "customer_id":      l_cust_ids,
    "loan_type":        l_types,
    "principal_amount": principal,
    "interest_rate_pct":int_rate,
    "tenure_months":    tenure,
    "outstanding_balance": outstanding,
    "emi_amount":       emi,
    "dpd":              dpd_list,
    "loan_status":      statuses,
    "credit_score":     credit_scores,
    "risk_grade":       risk_grades,
    "disbursal_date":   disbursal,
    "maturity_date":    maturity,
    "branch_code":      branch_codes,
})

# --- Loan_Payments (800) ---
pay_records = []
for loan_id, disb, emi_amt, status in zip(loan_ids, disbursal, emi, statuses):
    n_payments = random.randint(3, 24)
    for j in range(n_payments):
        due_date = disb + timedelta(days=30 * (j + 1))
        if due_date > date.today():
            break
        paid = emi_amt if status in ["Active", "Closed"] else (emi_amt if random.random() > 0.3 else round(emi_amt * random.uniform(0, 0.9), 2))
        delay = 0 if status == "Active" else random.randint(0, 60)
        pay_records.append({
            "payment_id":    f"PAY{len(pay_records)+1:06d}",
            "loan_id":       loan_id,
            "due_date":      due_date,
            "payment_date":  due_date + timedelta(days=delay),
            "scheduled_emi": emi_amt,
            "amount_paid":   paid,
            "payment_status": "On-Time" if delay == 0 else "Late" if paid >= emi_amt * 0.9 else "Partial",
            "delay_days":    delay,
            "payment_mode":  random.choice(["NEFT", "Auto-Debit", "Cheque", "UPI", "RTGS"]),
        })
    if len(pay_records) > 900:
        break

loan_payments = pd.DataFrame(pay_records[:900])

# --- Credit_Bureau (200) ---
credit_bureau = pd.DataFrame({
    "customer_id":         cust_ids,
    "bureau_score":        [random.randint(300, 900) for _ in range(200)],
    "active_loans_count":  [random.randint(0, 8) for _ in range(200)],
    "total_credit_limit":  [round(random.uniform(100000, 5000000), -3) for _ in range(200)],
    "credit_utilization_pct": [round(random.uniform(5, 95), 1) for _ in range(200)],
    "enquiries_last_6m":   [random.randint(0, 12) for _ in range(200)],
    "defaults_history":    [random.randint(0, 3) for _ in range(200)],
    "oldest_account_yrs":  [round(random.uniform(0.5, 20), 1) for _ in range(200)],
    "dpd_30_count":        [random.randint(0, 10) for _ in range(200)],
    "dpd_90_count":        [random.randint(0, 5) for _ in range(200)],
    "bureau_last_updated": [rand_date("2024-01-01", "2024-12-31") for _ in range(200)],
})

# --- Collateral (220) ---
COLL_TYPES = ["Residential Property", "Commercial Property", "Vehicle", "Fixed Deposit", "Gold", "Shares", "Machinery"]
collateral = pd.DataFrame({
    "collateral_id":    [f"COL{i:05d}" for i in range(1, 221)],
    "loan_id":          [random.choice(loan_ids) for _ in range(220)],
    "collateral_type":  [random.choice(COLL_TYPES) for _ in range(220)],
    "market_value":     [round(random.uniform(100000, 15000000), -3) for _ in range(220)],
    "forced_sale_value":[round(random.uniform(80000, 12000000), -3) for _ in range(220)],
    "ltv_ratio_pct":    [round(random.uniform(30, 90), 1) for _ in range(220)],
    "valuation_date":   [rand_date("2022-01-01", "2024-12-31") for _ in range(220)],
    "valuation_agency": [random.choice(["CRISIL", "ICRA", "CARE", "India Ratings", "Acuite"]) for _ in range(220)],
    "insurance_status": [random.choice(["Insured", "Insured", "Not Insured", "Expired"]) for _ in range(220)],
    "charge_type":      [random.choice(["Mortgage", "Hypothecation", "Pledge", "Assignment"]) for _ in range(220)],
})

path = OUT / "Banking_LoanPortfolio.xlsx"
with pd.ExcelWriter(path, engine="openpyxl") as w:
    customers.to_excel(w, index=False, sheet_name="Customers")
    loans.to_excel(w, index=False, sheet_name="Loans")
    loan_payments.to_excel(w, index=False, sheet_name="Loan_Payments")
    credit_bureau.to_excel(w, index=False, sheet_name="Credit_Bureau")
    collateral.to_excel(w, index=False, sheet_name="Collateral")
style_wb(path)
print(f"  Banking: {len(customers)} customers, {len(loans)} loans, {len(loan_payments)} payments, {len(credit_bureau)} bureau, {len(collateral)} collateral")


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Insurance_Claims.xlsx
#    Tabs: Policyholders → Policies → Claims → Fraud_Indicators
# ═══════════════════════════════════════════════════════════════════════════════
print("Generating Insurance_Claims.xlsx ...")

REGIONS   = ["North", "South", "East", "West", "Central"]
CLAIM_T   = ["Motor", "Health", "Property", "Life", "Travel", "Liability", "Marine"]
POL_TYPES = ["Comprehensive", "Third Party", "Individual Health", "Family Floater", "Group Health", "Term Life", "ULIP"]

# --- Policyholders (250) ---
ph_ids = [f"PH{i:05d}" for i in range(1, 251)]
policyholders = pd.DataFrame({
    "policyholder_id": ph_ids,
    "full_name":       [f"{random.choice(['James','Mary','John','Patricia','Robert','Jennifer','Michael','Linda','William','Barbara'])} {random.choice(['Smith','Johnson','Williams','Brown','Jones','Garcia','Miller','Davis','Wilson','Moore'])}" for _ in range(250)],
    "age":             [random.randint(21, 70) for _ in range(250)],
    "gender":          [random.choice(["Male","Female"]) for _ in range(250)],
    "region":          [random.choice(REGIONS) for _ in range(250)],
    "city":            [random.choice(["London","Manchester","Birmingham","Leeds","Glasgow","Liverpool","Bristol","Sheffield","Edinburgh","Cardiff"]) for _ in range(250)],
    "occupation":      [random.choice(["Employed","Self-Employed","Retired","Student","Business"]) for _ in range(250)],
    "annual_premium_total": [round(random.uniform(5000, 150000), 2) for _ in range(250)],
    "customer_since":  [rand_date("2010-01-01", "2022-12-31") for _ in range(250)],
    "risk_category":   [random.choice(["Low","Medium","Medium","High","Very High"]) for _ in range(250)],
})

# --- Policies (350) ---
pol_ids = [f"POL{i:06d}" for i in range(1, 351)]
pol_ph  = [random.choice(ph_ids) for _ in range(350)]
pol_start = [rand_date("2018-01-01", "2023-06-30") for _ in range(350)]
pol_end   = [s + timedelta(days=random.choice([365, 730, 1095])) for s in pol_start]
premiums  = [round(random.uniform(3000, 80000), 2) for _ in range(350)]
policies  = pd.DataFrame({
    "policy_id":       pol_ids,
    "policyholder_id": pol_ph,
    "policy_type":     [random.choice(POL_TYPES) for _ in range(350)],
    "sum_insured":     [round(random.uniform(100000, 10000000), -3) for _ in range(350)],
    "annual_premium":  premiums,
    "policy_start":    pol_start,
    "policy_end":      pol_end,
    "policy_status":   [random.choice(["Active","Active","Active","Lapsed","Cancelled"]) for _ in range(350)],
    "agent_code":      [f"AGT{random.randint(100,999)}" for _ in range(350)],
    "payment_freq":    [random.choice(["Annual","Semi-Annual","Quarterly","Monthly"]) for _ in range(350)],
    "no_claim_bonus_pct": [random.choice([0,5,10,15,20,25]) for _ in range(350)],
})

# --- Claims (450) ---
claim_ids  = [f"CLM{i:06d}" for i in range(1, 451)]
cl_pol_ids = [random.choice(pol_ids) for _ in range(450)]
cl_amounts = [round(random.uniform(1000, 800000), 2) for _ in range(450)]
cl_dates   = [rand_date("2021-01-01", "2024-12-31") for _ in range(450)]
proc_days  = [random.randint(1, 120) for _ in range(450)]
approved   = [round(a * random.uniform(0, 1.0), 2) if random.random() > 0.15 else 0 for a in cl_amounts]
prior_cl   = [random.randint(0, 9) for _ in range(450)]

claims = pd.DataFrame({
    "claim_id":           claim_ids,
    "policy_id":          cl_pol_ids,
    "claim_type":         [random.choice(CLAIM_T) for _ in range(450)],
    "claim_date":         cl_dates,
    "incident_date":      [d - timedelta(days=random.randint(0, 30)) for d in cl_dates],
    "claim_amount":       cl_amounts,
    "approved_amount":    approved,
    "claim_status":       [random.choice(["Approved","Approved","Rejected","Pending","Under Review","Settled"]) for _ in range(450)],
    "processing_days":    proc_days,
    "settlement_date":    [d + timedelta(days=p) for d, p in zip(cl_dates, proc_days)],
    "surveyor_id":        [f"SRV{random.randint(10,99)}" for _ in range(450)],
    "prior_claims_count": prior_cl,
    "region":             [random.choice(REGIONS) for _ in range(450)],
    "rejection_reason":   [random.choice(["Policy Lapsed","Exclusion Clause","Fraudulent","Duplicate","None","None","None","None"]) for _ in range(450)],
})

# --- Fraud_Indicators (180 suspicious claims) ---
suspicious = random.sample(claim_ids, 180)
fraud_scores = [round(random.uniform(0.3, 0.99), 3) for _ in range(180)]
fraud_df = pd.DataFrame({
    "fraud_id":          [f"FRD{i:05d}" for i in range(1, 181)],
    "claim_id":          suspicious,
    "fraud_score":       fraud_scores,
    "fraud_flag":        ["High Risk" if s > 0.7 else "Medium Risk" for s in fraud_scores],
    "indicator_1_multiple_claims": [random.choice([True, False]) for _ in range(180)],
    "indicator_2_late_reporting":  [random.choice([True, False]) for _ in range(180)],
    "indicator_3_inflated_amount": [random.choice([True, False]) for _ in range(180)],
    "indicator_4_known_associate": [random.choice([True, False]) for _ in range(180)],
    "indicator_5_prior_fraud":     [random.choice([True, False]) for _ in range(180)],
    "investigation_status": [random.choice(["Open","Closed","Referred to Police","Monitoring"]) for _ in range(180)],
    "analyst_id":        [f"ANL{random.randint(1,20):02d}" for _ in range(180)],
    "flagged_date":      [rand_date("2021-01-01", "2024-12-31") for _ in range(180)],
})

path = OUT / "Insurance_Claims.xlsx"
with pd.ExcelWriter(path, engine="openpyxl") as w:
    policyholders.to_excel(w, index=False, sheet_name="Policyholders")
    policies.to_excel(w, index=False, sheet_name="Policies")
    claims.to_excel(w, index=False, sheet_name="Claims")
    fraud_df.to_excel(w, index=False, sheet_name="Fraud_Indicators")
style_wb(path)
print(f"  Insurance: {len(policyholders)} policyholders, {len(policies)} policies, {len(claims)} claims, {len(fraud_df)} fraud flags")


# ═══════════════════════════════════════════════════════════════════════════════
# 3. Retail_Sales.xlsx
#    Tabs: Customers → Orders → Order_Items → Products → Returns
# ═══════════════════════════════════════════════════════════════════════════════
print("Generating Retail_Sales.xlsx ...")

CATEGORIES = ["Electronics", "Apparel", "Home & Kitchen", "Sports", "Beauty", "Books", "Toys", "Grocery"]
CHANNELS   = ["Online", "Store", "Marketplace", "B2B", "Social Commerce"]
REGIONS_R  = ["North", "South", "East", "West", "Central"]
SUB_CATS   = {
    "Electronics":    ["Mobile", "Laptop", "Tablet", "Headphones", "Camera", "Smartwatch"],
    "Apparel":        ["Shirts", "Jeans", "Dresses", "Shoes", "Jackets", "Sportswear"],
    "Home & Kitchen": ["Cookware", "Furniture", "Bedding", "Decor", "Appliances"],
    "Sports":         ["Footwear", "Equipment", "Clothing", "Nutrition", "Accessories"],
    "Beauty":         ["Skincare", "Haircare", "Makeup", "Fragrance", "Wellness"],
    "Books":          ["Fiction", "Non-Fiction", "Academic", "Children", "Business"],
    "Toys":           ["Action", "Educational", "Outdoor", "Board Games", "Dolls"],
    "Grocery":        ["Snacks", "Beverages", "Dairy", "Fresh Produce", "Packaged"],
}

# --- Customers (300) ---
rc_ids = [f"RC{i:05d}" for i in range(1, 301)]
retail_customers = pd.DataFrame({
    "customer_id":     rc_ids,
    "customer_name":   [f"Customer_{i}" for i in range(1, 301)],
    "age_group":       [random.choice(["18-25","26-35","36-45","46-55","55+"]) for _ in range(300)],
    "gender":          [random.choice(["Male","Female","Other"]) for _ in range(300)],
    "region":          [random.choice(REGIONS_R) for _ in range(300)],
    "city":            [random.choice(["New York","Los Angeles","Chicago","Houston","Phoenix","Philadelphia","San Antonio","San Diego","Dallas","Austin"]) for _ in range(300)],
    "channel_preference": [random.choice(CHANNELS) for _ in range(300)],
    "loyalty_tier":    [random.choice(["Bronze","Silver","Gold","Platinum"]) for _ in range(300)],
    "total_orders":    [random.randint(1, 50) for _ in range(300)],
    "lifetime_value":  [round(random.uniform(500, 150000), 2) for _ in range(300)],
    "acquisition_date":[rand_date("2019-01-01", "2023-12-31") for _ in range(300)],
    "nps_score":       [random.randint(0, 10) for _ in range(300)],
})

# --- Products (120) ---
prod_ids = [f"SKU{i:04d}" for i in range(1, 121)]
prod_list = []
for pid in prod_ids:
    cat = random.choice(CATEGORIES)
    sub = random.choice(SUB_CATS[cat])
    price = round(random.uniform(99, 79999), 2)
    prod_list.append({
        "product_id":   pid,
        "product_name": f"{sub} {random.choice(['Pro','Plus','Elite','Basic','Max','Lite','Ultra'])} {random.randint(100,999)}",
        "category":     cat,
        "sub_category": sub,
        "brand":        random.choice(["TechPro","FashionX","HomeComfort","SportElite","GlowBeauty","ReadMore","FunToys","FreshMart"]),
        "unit_price":   price,
        "cost_price":   round(price * random.uniform(0.35, 0.60), 2),
        "launch_date":  rand_date("2019-01-01", "2023-06-30"),
        "stock_units":  random.randint(0, 10000),
        "rating_avg":   round(random.uniform(2.5, 5.0), 1),
        "review_count": random.randint(0, 5000),
        "is_active":    random.choice([True, True, True, False]),
    })
products_df = pd.DataFrame(prod_list)

# --- Orders (700) ---
order_ids = [f"ORD{i:06d}" for i in range(1, 701)]
o_cust    = [random.choice(rc_ids) for _ in range(700)]
o_dates   = [rand_date("2022-01-01", "2024-12-31") for _ in range(700)]
# Seasonal boost Nov-Dec
o_dates   = [d if d.month not in [11, 12] else d.replace(day=random.randint(1, 28)) for d in o_dates]
o_revenue = [round(random.uniform(200, 250000), 2) for _ in range(700)]
o_disc    = [round(random.uniform(0, 0.40), 2) for _ in range(700)]
orders_df = pd.DataFrame({
    "order_id":        order_ids,
    "customer_id":     o_cust,
    "order_date":      o_dates,
    "channel":         [random.choice(CHANNELS) for _ in range(700)],
    "region":          [random.choice(REGIONS_R) for _ in range(700)],
    "gross_revenue":   o_revenue,
    "discount_pct":    o_disc,
    "net_revenue":     [round(r * (1 - d), 2) for r, d in zip(o_revenue, o_disc)],
    "order_status":    [random.choice(["Delivered","Delivered","Delivered","Shipped","Cancelled","Returned"]) for _ in range(700)],
    "payment_method":  [random.choice(["Credit Card","Debit Card","UPI","Net Banking","COD","Wallet"]) for _ in range(700)],
    "delivery_days":   [random.randint(1, 14) for _ in range(700)],
    "sale_month":      [d.strftime("%Y-%m") for d in o_dates],
})

# --- Order_Items (1400) ---
items = []
for oid in order_ids:
    n = random.randint(1, 4)
    for _ in range(n):
        p = random.choice(prod_list)
        qty = random.randint(1, 10)
        items.append({
            "item_id":    f"ITM{len(items)+1:07d}",
            "order_id":   oid,
            "product_id": p["product_id"],
            "category":   p["category"],
            "quantity":   qty,
            "unit_price": p["unit_price"],
            "line_total":  round(p["unit_price"] * qty, 2),
            "discount_pct": round(random.uniform(0, 0.35), 2),
        })
order_items_df = pd.DataFrame(items)

# --- Returns (150) ---
ret_orders = random.sample(order_ids, 150)
returns_df = pd.DataFrame({
    "return_id":     [f"RET{i:05d}" for i in range(1, 151)],
    "order_id":      ret_orders,
    "return_date":   [rand_date("2022-06-01", "2024-12-31") for _ in range(150)],
    "return_reason": [random.choice(["Defective","Wrong Item","Not as Described","Changed Mind","Size Issue","Duplicate Order"]) for _ in range(150)],
    "refund_amount": [round(random.uniform(200, 50000), 2) for _ in range(150)],
    "return_status": [random.choice(["Processed","Processed","Pending","Rejected"]) for _ in range(150)],
    "return_channel":[random.choice(["Online Portal","Store","Courier Pickup"]) for _ in range(150)],
    "resolution_days":[random.randint(1, 21) for _ in range(150)],
})

path = OUT / "Retail_Sales.xlsx"
with pd.ExcelWriter(path, engine="openpyxl") as w:
    retail_customers.to_excel(w, index=False, sheet_name="Customers")
    orders_df.to_excel(w, index=False, sheet_name="Orders")
    order_items_df.to_excel(w, index=False, sheet_name="Order_Items")
    products_df.to_excel(w, index=False, sheet_name="Products")
    returns_df.to_excel(w, index=False, sheet_name="Returns")
style_wb(path)
print(f"  Retail: {len(retail_customers)} customers, {len(orders_df)} orders, {len(order_items_df)} items, {len(products_df)} products, {len(returns_df)} returns")


# ═══════════════════════════════════════════════════════════════════════════════
# 4. HR_Workforce.xlsx
#    Tabs: Employees → Departments, Performance, Salary_History, Training
# ═══════════════════════════════════════════════════════════════════════════════
print("Generating HR_Workforce.xlsx ...")

DEPT_NAMES   = ["Engineering","Sales","HR","Finance","Operations","Marketing","Legal","Product","Support","Data Science"]
GRADES       = ["L1","L2","L3","L4","L5","L6","L7"]
LOCATIONS    = ["Bangalore","Mumbai","Delhi","Chennai","Hyderabad","Pune","Kolkata","Gurgaon"]
BASE_SALARY  = {"L1":350000,"L2":550000,"L3":850000,"L4":1300000,"L5":2000000,"L6":3000000,"L7":4500000}

# --- Departments (10) ---
dept_ids = [f"DEPT{i:02d}" for i in range(1, 11)]
departments_df = pd.DataFrame({
    "department_id":   dept_ids,
    "department_name": DEPT_NAMES,
    "department_head": [f"EMP{random.randint(1,30):05d}" for _ in range(10)],
    "location":        [random.choice(LOCATIONS) for _ in range(10)],
    "headcount_budget":[random.randint(20, 200) for _ in range(10)],
    "annual_budget":   [round(random.uniform(5000000, 100000000), -5) for _ in range(10)],
    "cost_center":     [f"CC{random.randint(1000,9999)}" for _ in range(10)],
    "established_year":[random.randint(2005, 2020) for _ in range(10)],
})

# --- Employees (500) ---
emp_ids = [f"EMP{i:05d}" for i in range(1, 501)]
e_depts = [random.choice(dept_ids) for _ in range(500)]
e_dept_names = [DEPT_NAMES[dept_ids.index(d)] for d in e_depts]
e_grades = [random.choices(GRADES, weights=[20,25,20,15,10,6,4])[0] for _ in range(500)]
e_salaries = [round(BASE_SALARY[g] * random.uniform(0.85, 1.25), -3) for g in e_grades]
e_tenure = [round(random.uniform(0.1, 20), 1) for _ in range(500)]
e_join  = [date.today() - timedelta(days=int(t * 365)) for t in e_tenure]
e_gender = [random.choice(["Male","Female","Non-Binary"]) for _ in range(500)]
# Pay gap
for i, g in enumerate(e_gender):
    if g == "Female" and random.random() > 0.55:
        e_salaries[i] = round(e_salaries[i] * random.uniform(0.78, 0.95), -3)

attrition_risk = []
for t, g, s in zip(e_tenure, e_grades, e_salaries):
    r = 0.08
    if t < 1: r += 0.28
    if g in ["L1","L2"] and t < 2: r += 0.15
    if s < 500000: r += 0.10
    attrition_risk.append(round(min(r + random.uniform(-0.05, 0.15), 0.95), 2))

JOB_MAP = {
    "Engineering":  ["Software Engineer","Senior Engineer","Tech Lead","Principal Engineer","Architect"],
    "Sales":        ["Sales Executive","Account Manager","Sales Lead","Regional Manager","VP Sales"],
    "HR":           ["HR Associate","HR Business Partner","HR Manager","HR Director","CHRO"],
    "Finance":      ["Financial Analyst","Senior Analyst","Finance Manager","Controller","CFO"],
    "Operations":   ["Ops Associate","Ops Manager","Sr Operations Manager","Director Ops","COO"],
    "Marketing":    ["Marketing Executive","Brand Manager","Digital Marketing Lead","Marketing Director","CMO"],
    "Legal":        ["Legal Associate","Legal Manager","Senior Counsel","General Counsel","CLO"],
    "Product":      ["Product Analyst","Product Manager","Senior PM","Group PM","VP Product"],
    "Support":      ["Support Agent","Support Lead","Support Manager","Head of Support","VP CX"],
    "Data Science": ["Data Analyst","Data Scientist","Senior Data Scientist","ML Engineer","Chief Data Officer"],
}

employees_df = pd.DataFrame({
    "employee_id":       emp_ids,
    "full_name":         [f"{random.choice(['Arjun','Priya','Rahul','Sneha','Vikram','Anjali','Rohan','Kavya','Aditya','Meghna'])} {random.choice(['Sharma','Patel','Gupta','Singh','Kumar','Nair','Reddy','Joshi','Iyer','Pillai'])}" for _ in range(500)],
    "department_id":     e_depts,
    "department":        e_dept_names,
    "job_title":         [random.choice(JOB_MAP[dn]) for dn in e_dept_names],
    "grade":             e_grades,
    "location":          [random.choice(LOCATIONS) for _ in range(500)],
    "gender":            e_gender,
    "age":               [random.randint(22, 58) for _ in range(500)],
    "education":         [random.choice(["Bachelor's","Master's","MBA","PhD","Diploma"]) for _ in range(500)],
    "join_date":         e_join,
    "tenure_years":      e_tenure,
    "annual_salary":     e_salaries,
    "manager_id":        [f"EMP{random.randint(1,50):05d}" for _ in range(500)],
    "wfh_days_per_week": [random.randint(0, 5) for _ in range(500)],
    "attrition_risk_score": attrition_risk,
    "employment_status": [random.choice(["Active","Active","Active","Active","On Leave","Probation"]) for _ in range(500)],
})

# --- Performance (1500 = 500 employees x 3 years) ---
perf_records = []
for eid, dept in zip(emp_ids, e_dept_names):
    for yr in [2022, 2023, 2024]:
        rating = random.choices([1,2,3,4,5], weights=[5,10,40,30,15])[0]
        perf_records.append({
            "review_id":         f"REV{len(perf_records)+1:06d}",
            "employee_id":       eid,
            "department":        dept,
            "review_year":       yr,
            "performance_rating":rating,
            "rating_label":      {1:"Needs Improvement",2:"Below Expectations",3:"Meets Expectations",4:"Exceeds Expectations",5:"Outstanding"}[rating],
            "goals_met_pct":     round(random.uniform(40, 100), 1),
            "manager_score":     round(random.uniform(1, 5), 1),
            "peer_score":        round(random.uniform(1, 5), 1),
            "training_hours":    random.randint(0, 80),
            "promoted":          random.choice([True, False, False, False, False]),
            "bonus_pct":         round(random.uniform(8, 25), 1) if rating >= 4 else round(random.uniform(0, 7), 1),
        })
performance_df = pd.DataFrame(perf_records)

# --- Salary_History (800) ---
sal_records = []
for eid, curr_sal in zip(random.sample(emp_ids, 400), random.sample(e_salaries, 400)):
    n_changes = random.randint(1, 3)
    sal = round(curr_sal * random.uniform(0.6, 0.9), -3)
    for _ in range(n_changes):
        hike = round(random.uniform(0.05, 0.25), 3)
        sal_records.append({
            "salary_record_id": f"SAL{len(sal_records)+1:06d}",
            "employee_id":      eid,
            "effective_date":   rand_date("2018-01-01", "2024-01-01"),
            "previous_salary":  sal,
            "new_salary":       round(sal * (1 + hike), -3),
            "hike_pct":         round(hike * 100, 1),
            "revision_type":    random.choice(["Annual Increment","Promotion","Market Correction","Performance Bonus"]),
            "approved_by":      f"EMP{random.randint(1,30):05d}",
        })
        sal = round(sal * (1 + hike), -3)
salary_history_df = pd.DataFrame(sal_records)

# --- Training (600) ---
training_df = pd.DataFrame({
    "training_id":    [f"TRN{i:05d}" for i in range(1, 601)],
    "employee_id":    [random.choice(emp_ids) for _ in range(600)],
    "course_name":    [random.choice(["Python for Data","Leadership 101","AWS Certified","Agile & Scrum","Excel Mastery","Communication Skills","Project Management","AI/ML Basics","Sales Excellence","Finance for Non-Finance"]) for _ in range(600)],
    "category":       [random.choice(["Technical","Leadership","Compliance","Soft Skills","Domain"]) for _ in range(600)],
    "completion_date":[rand_date("2021-01-01", "2024-12-31") for _ in range(600)],
    "hours":          [random.randint(4, 40) for _ in range(600)],
    "score_pct":      [round(random.uniform(50, 100), 1) for _ in range(600)],
    "status":         [random.choice(["Completed","Completed","In Progress","Not Started"]) for _ in range(600)],
    "training_mode":  [random.choice(["Online","Classroom","Workshop","On-the-Job"]) for _ in range(600)],
    "cost":           [round(random.uniform(500, 50000), 2) for _ in range(600)],
})

path = OUT / "HR_Workforce.xlsx"
with pd.ExcelWriter(path, engine="openpyxl") as w:
    employees_df.to_excel(w, index=False, sheet_name="Employees")
    departments_df.to_excel(w, index=False, sheet_name="Departments")
    performance_df.to_excel(w, index=False, sheet_name="Performance")
    salary_history_df.to_excel(w, index=False, sheet_name="Salary_History")
    training_df.to_excel(w, index=False, sheet_name="Training")
style_wb(path)
print(f"  HR: {len(employees_df)} employees, {len(departments_df)} depts, {len(performance_df)} reviews, {len(salary_history_df)} salary records, {len(training_df)} trainings")


print()
print("All 4 use-case datasets saved to:", OUT)
print()
for f in sorted(OUT.glob("*.xlsx")):
    size_kb = f.stat().st_size // 1024
    import openpyxl
    wb = openpyxl.load_workbook(f, read_only=True)
    tabs = ", ".join(wb.sheetnames)
    print(f"  {f.name:<38} {size_kb:>4} KB   [{tabs}]")
